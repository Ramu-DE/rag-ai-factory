# -*- coding: utf-8 -*-
"""
Build Mindee_Capabilities.xlsx
================================
Comprehensive reference workbook covering every Mindee platform capability
researched from https://docs.mindee.com and https://app.mindee.com

Run:  python build_mindee_capabilities_xlsx.py
Out:  docs/Mindee_Capabilities.xlsx
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── palette ──────────────────────────────────────────────────────────────────
C = {
    "navy":    "1A2F5A", "blue":    "1565C0", "teal":    "00695C",
    "green":   "2E7D32", "purple":  "6A1B9A", "orange":  "E65100",
    "red":     "B71C1C", "brown":   "4E342E", "indigo":  "283593",
    "cyan":    "006064", "white":   "FFFFFF", "lgrey":   "F5F5F5",
    "mgrey":   "E0E0E0", "dgrey":   "424242", "black":   "000000",
    "sky":     "E3F2FD", "mint":    "E8F5E9", "lav":     "EDE7F6",
    "peach":   "FBE9E7", "lemon":   "FFFDE7", "ice":     "E0F7FA",
    "pink":    "FCE4EC", "sand":    "FFF8E1",
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, sz=10, color="000000", italic=False, name="Calibri"):
    return Font(bold=bold, size=sz, color=color, italic=italic, name=name)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)
def hdr_border():
    t = Side(style="thin",   color="BDBDBD")
    b = Side(style="medium", color="757575")
    return Border(left=t, right=t, top=t, bottom=b)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, title, sub, bg, r=1):
    ws.row_dimensions[r].height   = 32
    ws.row_dimensions[r+1].height = 16
    c1 = ws.cell(r,   1, title)
    c1.font      = font(True, 16, "FFFFFF")
    c1.fill      = fill(bg)
    c1.alignment = align("left", "center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    c2 = ws.cell(r+1, 1, sub)
    c2.font      = font(False, 10, C["dgrey"], True)
    c2.fill      = fill(C["lgrey"])
    c2.alignment = align("left", "center")
    ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=12)
    return r + 3

def hdr_row(ws, r, cols, bgs, fg="FFFFFF"):
    for ci, (txt, bg) in enumerate(zip(cols, bgs), 1):
        c = ws.cell(r, ci, txt)
        c.font      = font(True, 10, fg)
        c.fill      = fill(bg)
        c.alignment = align("center", "center", True)
        c.border    = hdr_border()

def data_row(ws, r, vals, bg=None, bold=False, wrap=True, center_cols=None, start=1):
    for ci, v in enumerate(vals, start):
        cell = ws.cell(r, ci, v)
        if bg: cell.fill = fill(bg)
        cell.font      = font(bold, 10)
        cell.alignment = align(
            "center" if (center_cols and ci in center_cols) else "left",
            "center", wrap
        )
        cell.border = border()

def freeze(ws, cell="A2"): ws.freeze_panes = cell

def section_hdr(ws, r, text, bg, span=12):
    c = ws.cell(r, 1, text)
    c.font      = font(True, 11, "FFFFFF")
    c.fill      = fill(bg)
    c.alignment = align("left", "center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    ws.row_dimensions[r].height = 20
    return r + 1

# ═══════════════════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════════════════

PLATFORM_OVERVIEW = [
    # (Category, Feature, What It Does, Key Benefit, Availability)
    ("Core", "Extraction Models",
     "OCR + AI extracts structured fields from any document type — invoices, IDs, contracts, healthcare cards, resumes, and more.",
     "Converts unstructured PDFs/images into clean JSON with confidence scores.", "All plans"),
    ("Core", "Custom Model Builder",
     "Create bespoke extraction models by uploading a sample doc or describing the document type; AI Agent auto-generates a data schema.",
     "No ML expertise needed — AI scaffolds the schema in seconds.", "All plans"),
    ("Core", "Catalog Templates",
     "14 pre-built extraction model templates (Invoice 26 fields, Receipt 17, Financial Document 32, Passport 12, Resume 13, International ID 13, Driver License 14, Business Card 8, US Healthcare Card 14, Boarding Pass 7, Bill of Lading 11, Nutrition Facts 15, Bank Statement 16, Payslip 20) ready to use or customise.",
     "Zero setup time for common document types — exact field counts visible on platform.", "All plans"),
    ("Utility", "Crop",
     "Identify the borders of documents on each page and match each one to a category. Returns bounding polygon + user-defined class per detected region.",
     "Handles batch scans where multiple docs appear on one page.", "Available"),
    ("Utility", "Split",
     "Break a multi-page source file into separate documents and associate a class to each one. Returns page ranges + user-defined category per segment.",
     "Automates processing of mixed-document bundles.", "Available"),
    ("Utility", "Classify",
     "Automatically sort any image or scanned document based on your categories. Returns predicted class + confidence score.",
     "Routes documents to correct workflows automatically.", "Available"),
    ("Utility", "OCR (Raw Text)",
     "Extract raw text from any image or scanned document with high precision. Returns full page text + word-level polygons.",
     "Feeds downstream NLP/RAG pipelines with positioned text.", "Available"),
    ("Optional", "Confidence Score + Accuracy Boost",
     "Ensemble of models; consensus-based scoring returns Certain/High/Medium/Low per field with colour-coded automation signal.",
     "Safe full-automation for high-confidence fields; human review for Low.", "Add-on"),
    ("Optional", "Polygons / Bounding Boxes",
     "Normalised [0-1] polygon coordinates for every extracted field; clockwise point list, centroid helper included.",
     "Visual overlay for end-user review UIs; enables field localisation.", "Add-on"),
    ("Optional", "Raw Text Full OCR",
     "Adds complete page text alongside structured fields in every API response.",
     "No second API call needed for full OCR alongside extraction.", "Add-on"),
    ("Optional", "Continuous Learning / RAG",
     "Provide correction examples; model improves durably using Retrieval-Augmented Generation feedback loop.",
     "Low-confidence fields become training signals; accuracy compounds over time.", "Add-on"),
    ("Integration", "SDKs (6 languages)",
     "Official client libraries: Python, Node.js, Java, Ruby, PHP, .NET (C#). Handle file upload, polling, error handling automatically.",
     "Synchronous-style API calls despite async backend.", "All plans"),
    ("Integration", "No-Code (Make, n8n, Zapier)",
     "Native app connectors for Make.com, n8n, and Zapier. Connect Mindee V2 to 5,000+ apps without writing code.",
     "Document automation in minutes for non-developers.", "All plans"),
    ("Integration", "Webhooks",
     "POST inference results to your server on completion; HMAC signing secret for payload verification; static IPs on enterprise.",
     "Best for high-volume production — no polling overhead.", "All plans"),
    ("Integration", "Polling",
     "Submit to /enqueue, get job_id, poll /jobs/{id}; wait 3s first then 1s intervals; SDK abstracts this entirely.",
     "Good for local testing and low-volume scenarios.", "All plans"),
    ("Platform", "Live Test UI",
     "Upload a document in the browser, see extraction results + confidence overlays + JSON inspector in real time.",
     "Visual validation without writing a single line of code.", "All plans"),
    ("Platform", "Data Residency & Compliance",
     "GDPR compliant; SOC 2 Type II + ISO 27001 ready; EU/US data residency selection; TLS encryption in transit.",
     "Enterprise-grade data governance out of the box.", "Enterprise"),
    ("Platform", "Role-Based Access Control",
     "Organisations with teams; granular model/API key permissions per member.",
     "Secure multi-team collaboration.", "Team/Enterprise"),
    ("API", "Rate Limits",
     "POST: 200 req/min; GET polling: 1,200 req/min; HTTP 429 on breach; enterprise customisation available.",
     "Predictable throughput planning.", "All plans"),
    ("API", "File Limits",
     "Max 100 MB per file; max 200 pages; formats: PDF (all variants, unencrypted), JPEG, PNG, WebP, TIFF, HEIC, HEIF.",
     "Covers virtually all real-world document formats.", "All plans"),
]

MODELS_ALL = [
    # Field counts verified from app.mindee.com/create-model UI (shown field + "See N more")
    # (Category, Model Name, Endpoint Slug, Doc Types, Field Count, Key Fields Summary)

    # ── Pre-built: Finance ────────────────────────────────────────────────────────
    ("Finance", "Invoice",
     "mindee/invoices/v4",
     "Invoice, bill, purchase order",
     "26",   # UI: Supplier Name + See 25 more
     "supplier_name, invoice_number, date, due_date, total_amount, total_net, total_tax, "
     "line_items (description/qty/unit_price/total_price/product_code/tax), taxes (rate/base/amount), "
     "supplier_address, customer_address, billing_address, shipping_address, "
     "supplier_payment_details (IBAN/SWIFT/account/routing), customer_company_registration, "
     "supplier_company_registration, reference_numbers, po_number, document_type, locale (language/country/currency)"),
    ("Finance", "Receipt",
     "mindee/receipt_ocrs/v5",
     "Expense receipt, credit card receipt",
     "17",   # UI: Supplier Name + See 16 more
     "supplier_name, supplier_address, supplier_phone_number, receipt_number, date, time, "
     "document_type, total_amount, total_net, total_tax, tips_gratuity, taxes, "
     "line_items, purchase_category, purchase_subcategory, locale"),
    ("Finance", "Financial Document",
     "mindee/financial_document/v1",
     "Invoice, receipt, credit note, payslip, quote, purchase order, statement",
     "32",   # UI: Supplier Name + See 31 more
     "Auto-detects document sub-type. Union of Invoice + Receipt fields: "
     "supplier_name/phone/email/website, customer_name/id, invoice_number, document_number, "
     "date, due_date, payment_date, total_amount/net/tax, tip, line_items, taxes, "
     "customer_address, supplier_address, billing_address, shipping_address, "
     "customer_company_registration, supplier_company_registration, "
     "supplier_payment_details, po_number, reference_numbers, category, subcategory, locale"),
    ("Finance", "Bank Statement",
     "mindee/bank_account_details/v2",
     "Bank statements (all major banks)",
     "16",
     "account_holder_names, account_number, account_type, currency, "
     "statement_period_start, statement_period_end, statement_date, "
     "beginning_balance, ending_balance, total_credits, total_debits, "
     "daily_balances (date+amount list), transactions (date+description+amount list), "
     "bank_name, branch_code, bank_address, account_holder_address"),
    ("Finance", "Payslip (FR)",
     "mindee/payslip_fra/v3",
     "French payslips",
     "20",
     "employee_name, employee_id, employee_address, company_name, company_address, "
     "pay_period, gross_salary, net_salary, total_deductions, "
     "pay_components (name/amount/base), social_security_number, payment_date"),

    # ── Pre-built: Identity ───────────────────────────────────────────────────────
    ("Identity", "Passport",
     "mindee/passports/v1",
     "International passports (all countries)",
     "12",   # UI: Given Names + See 11 more
     "given_names, surnames, date_of_birth, place_of_birth, passport_number, "
     "issuing_country, nationality, date_of_issue, date_of_expiry, sex, "
     "mrz_line_1, mrz_line_2"),
    ("Identity", "International ID",
     "mindee/international_id/v1",
     "National IDs, residence permits, passports, driver licences",
     "13",   # UI: Given Names + See 12 more
     "given_names, surnames, date_of_birth, place_of_birth, nationality, sex, "
     "document_number, date_of_issue, date_of_expiry, authority, document_type, "
     "address (street/city/state/postal_code/country), mrz (line_1/line_2/line_3)"),
    ("Identity", "Driver's License",
     "mindee/us_driver_license/v1",
     "Driver licences worldwide",
     "14",   # UI: First Name + See 13 more
     "first_name, last_name, date_of_birth, place_of_birth, sex, "
     "document_id, issued_date, expiry_date, country_code, issuing_authority, "
     "mrz, category (EU driving categories A/B/C/D), street, city, state, postal_code"),
    ("Identity", "Business Card",
     "mindee/business_card/v1",
     "Printed and digital business cards",
     "8",
     "name, job_title, company, phone_numbers (array), "
     "emails (array), website, address, social_media (array)"),

    # ── Pre-built: Healthcare ─────────────────────────────────────────────────────
    ("Healthcare", "US Healthcare Card",
     "mindee/us_healthcare_cards/v1",
     "US health insurance member cards",
     "14",
     "company_name, plan_name, member_name, member_id, issuer_80840, "
     "group_number, payer_id, dependents (array), "
     "rx_bin, rx_id, rx_grp, rx_pcn, "
     "copayments (service_name + service_fees array), enrollment_date"),

    # ── Pre-built: HR ─────────────────────────────────────────────────────────────
    ("HR", "Resume / CV",
     "mindee/resume_and_cv/v1",
     "Resumes and CVs (all formats)",
     "13",   # UI: Name + Address + See 11 more
     "name, address, phone_number, email, linkedin_profile, "
     "education (school_name/degree/dates_attended/gpa/relevant_coursework array), "
     "professional_history (company_name/job_title/dates_employed/responsibilities array), "
     "skills (array), languages (name+proficiency array), "
     "projects (name/description/technologies array), "
     "awards_certifications (name+date array), summary, candidate_photo (object detection)"),

    # ── Pre-built: Logistics ──────────────────────────────────────────────────────
    ("Logistics", "Boarding Pass",
     "mindee/barcode_reader/v1",
     "Airline boarding passes (print + digital)",
     "7",
     "passenger_name, flight_number, departure_date, boarding_time, "
     "departure_airport (name+IATA), destination_airport (name+IATA), "
     "booking_reference (PNR)"),
    ("Logistics", "Bill of Lading",
     "mindee/bill_of_lading/v1",
     "International shipping bills of lading",
     "11",
     "bill_of_lading_number, shipper (name/address/phone/email), "
     "consignee (name/address/phone/email), notify_party (contact info), "
     "carrier (name/professional_number/SCAC), "
     "carrier_items (description/quantity/gross_weight/weight_unit/measurement/measurement_unit), "
     "port_of_loading, port_of_discharge, place_of_delivery, issue_date, departure_date"),

    # ── Pre-built: Labels ─────────────────────────────────────────────────────────
    ("Labels", "Nutrition Facts",
     "mindee/nutrition_facts/v1",
     "US/EU nutrition facts labels",
     "15",
     "serving_size, servings_per_container, calories, "
     "total_fat/saturated_fat/trans_fat/cholesterol/sodium (amount+daily_value), "
     "total_carbohydrate/dietary_fiber/total_sugars/added_sugars (amount+daily_value), "
     "protein, vitamins (name+amount+daily_value array), ingredients"),

    # ── Document Utilities (custom) ───────────────────────────────────────────────
    ("Utility — Crop",
     "Crop",
     "custom model ID",
     "Any page containing multiple documents (batch scan, clustered photo)",
     "—",
     "Identify the borders of documents on each page and match each one to a category. "
     "Returns per region: bounding_box (normalised polygon), category (user-defined class), page_index."),
    ("Utility — Split",
     "Split",
     "custom model ID",
     "Multi-page PDF/image containing several distinct documents",
     "—",
     "Break a multi-page source file into separate documents and associate a class to each one. "
     "Returns per segment: page_start, page_end, category (user-defined class)."),
    ("Utility — OCR",
     "OCR",
     "custom model ID",
     "Any image or scanned document, almost all languages",
     "—",
     "Extract raw text from any image or scanned document with high precision. "
     "Returns per page: full_text (string) + words (text + polygon + page_index array)."),
    ("Utility — Classify",
     "Classify",
     "custom model ID",
     "Any single or multi-page document",
     "—",
     "Automatically sort any image or scanned document based on your categories. "
     "Returns: predicted_class (one of your defined categories) + confidence score."),
]

FIELD_TYPES = [
    # (Field Type, Description, Example Field, Output Format, Array Support, Notes)
    ("Text",       "Sequence of characters; general string value",
     "supplier_name, invoice_number, address",
     "string or null",  "Yes", "Most common type; handles multilingual text"),
    ("Number",     "Integer or floating-point decimal value",
     "total_amount, quantity, unit_price",
     "float or null",   "Yes", "Python returns float; Java uses Double"),
    ("Date",       "Calendar date; optionally with time",
     "date, due_date, date_of_birth",
     "YYYY-MM-DD or YYYY-MM-DD HH:mm:ss or null",
     "Yes", "Time component optional"),
    ("Boolean",    "True/false value; name should start with 'is' or 'has'",
     "is_signed, has_barcode, is_electronic",
     "true / false / null",
     "Yes", "Best for checkboxes, flags, and binary indicators"),
    ("Classification", "Categorises value into one of a predefined set of classes",
     "document_type, purchase_category, sex",
     "enum string or null",
     "No",  "Classes defined at schema time; null if no match"),
    ("Nested Object", "Groups multiple related sub-fields under one parent",
     "locale (language+country+currency), address (street+city+state+zip+country)",
     "object with named sub-fields",
     "Yes (becomes array of objects)",
     "Single level of nesting only; up to 25 fields recommended total"),
    ("Object Detection", "Locates a visual element (logo, signature, barcode image)",
     "candidate_photo, barcode_image",
     "polygon coordinates + page_index",
     "Yes", "Returns position not text; useful for visual verification"),
    ("Barcode",    "Detects and decodes 1D and 2D barcodes",
     "barcode, qr_code",
     "decoded string value + polygon",
     "Yes", "Handles QR, Code128, EAN, PDF417 and others"),
]

FIELD_RULES = [
    # (Rule, Detail)
    ("Name format",         "Lowercase a-z, digits 0-9, underscores only. No leading/trailing underscores. No accents or special chars."),
    ("Max fields",          "Recommended maximum 25 fields per schema. Beyond 25, response latency increases noticeably."),
    ("Array fields",        "Any base field type can be made an array by enabling 'Multiple items can be extracted'. Duplicate filtering available."),
    ("Null values",         "All fields can return null — the API never errors on missing data; confidence will be Low."),
    ("Global guidelines",   "You can add schema-level extraction instructions that apply across all fields (e.g. 'Use ISO date format')."),
    ("Field guidelines",    "Per-field optional 'Guidelines' text gives the model specific extraction instructions for that field."),
    ("Field description",   "Per-field optional 'Description' provides context about what the field represents — improves AI accuracy."),
    ("Multilingual support","Field titles, names, descriptions, and guidelines can all be in any language the model supports."),
    ("Nesting depth",       "Only one level of nesting supported. Nested objects cannot themselves contain nested objects."),
]

CONFIDENCE_LEVELS = [
    # (Level, Color, Meaning, Recommended Action, Technical Detail)
    ("Certain", "Blue",
     "Full human-level confidence; multiple models agree completely",
     "Safe for full automation — route directly to ERP/CRM without review",
     "Highest consensus across ensemble; structural + semantic coherence confirmed by arbitration model"),
    ("High", "Green",
     "Strong consensus; prediction very likely accurate",
     "Auto-process; optional spot-check on high-value fields",
     "High agreement across models; minor formatting variations tolerated"),
    ("Medium", "Orange",
     "Some confidence; document quality or format may have impacted accuracy",
     "Optional human review; flag for QA queue in high-stakes workflows",
     "Models partially agree; document noise, stamps, or unusual layouts detected"),
    ("Low", "Red",
     "Extraction uncertain or likely incorrect",
     "Route to manual review queue; use as training signal for RAG feedback",
     "Significant disagreement between models; becomes feedback for continuous learning"),
]

API_ENDPOINTS = [
    # (Method, Path, Purpose, Auth, Request, Response)
    ("POST", "/v2/inferences/enqueue",
     "Submit a file for processing (all model types)",
     "Authorization: Token <api_key>",
     "multipart/form-data: file (binary) + model_id (string) + optional feature flags",
     "HTTP 202 + {job_id, status:'enqueued'}"),
    ("GET", "/v2/jobs/{job_id}",
     "Poll for processing status",
     "Authorization: Token <api_key>",
     "—",
     "HTTP 200 (still processing) | HTTP 302 with result_url when complete"),
    ("GET", "/v2/inferences/{inference_id}",
     "Retrieve completed inference result",
     "Authorization: Token <api_key>",
     "—",
     "HTTP 200 + full inference JSON: fields, confidence, polygons, raw_text"),
    ("POST", "(webhook URL)",
     "Mindee POSTs inference result to your server on completion",
     "HMAC signing secret in X-Mindee-Signature header",
     "Identical payload to GET inference result",
     "Your server must return HTTP 2xx; no redirects followed"),
    ("GET", "/v2/inferences/{inference_id}",
     "Fallback retrieval if webhook missed",
     "Authorization: Token <api_key>",
     "—",
     "Same as polling result; available for at least 1 hour after processing"),
]

RESPONSE_STRUCTURE = [
    # (Path, Type, Description)
    ("response.inference.result.fields",              "dict",         "Top-level map of field_name → field_object"),
    ("fields.<name>.value",                           "any / null",   "Extracted value (SimpleField: string, number, bool, date)"),
    ("fields.<name>.confidence",                      "enum / null",  "Certain | High | Medium | Low (null if feature off)"),
    ("fields.<name>.locations",                       "list / null",  "List of FieldLocation objects (null if feature off)"),
    ("fields.<name>.locations[i].page",               "int",          "Zero-based page index where field was found"),
    ("fields.<name>.locations[i].polygon",            "list[Point]",  "Clockwise list of normalised [x,y] points (0.0–1.0); index 0 = top-left"),
    ("fields.<name>.locations[i].polygon.centroid",   "Point",        "Geometric centroid helper function"),
    ("fields.<name>.fields",                          "dict",         "Sub-fields map (ObjectField only)"),
    ("fields.<name>.items",                           "list",         "Array of SimpleField or ObjectField (ListField only)"),
    ("response.inference.result.options.raw_text",    "list[str]",    "Full page text per page (when raw_text feature enabled)"),
    ("response.inference.result.options.words",       "list[Word]",   "Per-page word list with text + polygon (when raw_text feature enabled)"),
    ("response.inference.id",                         "str",          "Unique inference ID for retrieval"),
    ("response.inference.created_at",                 "datetime",     "ISO 8601 timestamp of inference creation"),
    ("response.inference.model_id",                   "str",          "ID of the model that processed the document"),
    ("response.inference.file.name",                  "str",          "Original filename submitted"),
    ("response.inference.file.page_count",            "int",          "Number of pages in the processed file"),
]

SDK_DETAILS = [
    # (Language, Install, Key Methods, Notes)
    ("Python",  "pip install mindee",
     "client = Client(api_key='...')\nresult = client.enqueue_and_parse(model, file)\nresult.inference.result.fields['supplier_name'].value",
     "Auto-polls until complete. Supports sync + async. Full type hints."),
    ("Node.js / TS",
     "npm install mindee",
     "const client = new Client({apiKey:'...'})\nconst result = await client.enqueueAndParse(ProductV2, {path:'file.pdf'})\nresult.inference.result.fields.supplierName.stringValue",
     "Typed accessors (.stringValue, .numberValue, .booleanValue). Promise-based."),
    ("Java",    "Maven / Gradle: com.mindee:mindee-api-java",
     "MindeeClient client = new MindeeClient(apiKey);\nPredictResponse<ProductV2> result = client.enqueueAndParse(ProductV2.class, file);\nresult.getDocument().getInference().getResult().getFields()",
     "Explicit type declarations required. Numbers return as Double."),
    ("Ruby",    "gem install mindee",
     "client = Mindee::Client.new(api_key:'...')\nresult = client.enqueue_and_parse(Mindee::Product::InvoiceV4, 'file.pdf')\nresult.document.inference.result.fields[:supplier_name].value",
     "Idiomatic Ruby API. Symbol-based field access."),
    ("PHP",     "composer require mindee/mindee",
     "$client = new Client('api_key');\n$result = $client->enqueueAndParse(InvoiceV4::class, 'file.pdf');\n$result->document->inference->result->fields['supplier_name']->value",
     "PSR-4 compliant. Namespace: Mindee\\Product."),
    (".NET / C#", "NuGet: Mindee",
     "var client = new MindeeClient(apiKey);\nvar result = await client.EnqueueAndParseAsync<InvoiceV4>(file);\nresult.Document.Inference.Result.Fields[\"supplier_name\"].Value",
     "Async/await first-class. Dynamic field resolution."),
    ("CLI",     "pip install mindee  (Python SDK includes CLI)",
     "mindee parse mindee/invoices/v4 invoice.pdf --api-key md_xxx",
     "Good for quick local tests without writing code. JSON output to stdout."),
]

NOCODE_INTEGRATIONS = [
    # (Platform, Setup Steps, Trigger, Action, Output Destinations, Notes)
    ("Make.com",
     "1. Add 'Mindee V2 verified' module\n2. Action: 'Extract Document Data'\n3. Create connection (name: MindeeV2-<keyname>, API key)\n4. Search model by name\n5. Link file input module (do NOT fill 'File' manually)",
     "Email attachment, Google Drive, OneDrive, Dropbox, Slack, S3, FTP, HTTP URL",
     "Extract Document Data → all fields returned as mapped variables",
     "Google Sheets, Salesforce, HubSpot, Notion, Slack, CSV export",
     "Use 'Mindee V2 verified' NOT community apps (which only support V1). File auto-connects from input module."),
    ("n8n",
     "1. Add Mindee node\n2. Configure API key credential\n3. Select document type / model\n4. Connect file source node",
     "Webhook trigger, file watcher, email reader, FTP",
     "Parse document → structured JSON output per field",
     "PostgreSQL, MySQL, HTTP request, email, spreadsheet",
     "Open-source option. Self-hostable. HTTP workflow nodes for custom API calls."),
    ("Zapier",
     "1. Search 'Mindee' in Zapier app directory\n2. Select trigger (new file in Drive, email, etc.)\n3. Add Mindee 'Parse Document' action\n4. Map extracted fields to next step",
     "Gmail attachment, Google Drive, Dropbox, Box",
     "Parse Document → field values as Zap variables",
     "Google Sheets, Airtable, Slack, Salesforce, QuickBooks",
     "Automation setup in seconds. 5,000+ connected apps."),
]

WEBHOOK_DETAILS = [
    ("Payload Format",        "Identical JSON to GET /v2/inferences/{id} response — full inference object"),
    ("Trigger Event",         "Fires once per inference on successful processing completion"),
    ("Security",              "HMAC signing secret available; validate X-Mindee-Signature header before processing payload"),
    ("Auth on URL",           "Credentials allowed as URL query params (?token=xxx or ?username=u&password=p); header auth NOT supported"),
    ("TLS Required",          "Endpoint must use HTTPS. HTTP not accepted."),
    ("Public URL Required",   "Endpoint must be publicly accessible — no VPN, localhost, or private network URLs"),
    ("Redirects",             "Mindee does NOT follow HTTP 3xx redirects; return 2xx directly"),
    ("Success Response",      "Your server must return HTTP 2xx; any other code = webhook delivery failure"),
    ("Retry / Storage",       "Failed inference available on server for minimum 1 hour; poll /jobs/{id} or /inferences/{id} as fallback"),
    ("Static IP",             "Available to enterprise customers on request (for firewall whitelisting)"),
    ("Multiple Endpoints",    "Create separate endpoints per environment (dev / staging / prod) in platform settings"),
    ("Job ID",                "Store the job_id from /enqueue response as retrieval fallback if webhook missed"),
]

RATE_LIMITS = [
    ("POST requests",         "200 per minute",  "HTTP 429",   "Submission and inference creation"),
    ("GET polling requests",  "1200 per minute", "HTTP 429",   "Job status and result retrieval"),
    ("File size limit",       "100 MB",          "HTTP 413",   "Per submitted file"),
    ("Page count limit",      "200 pages",       "HTTP 422",   "Per submitted file"),
    ("Schema field limit",    "25 fields (recommended)", "No hard error — latency increases", "Data schema per model"),
    ("Enterprise",            "Customisable",    "Contact sales", "Organisation-level rate limit increases available"),
]

FILE_FORMATS = [
    ("PDF",  "All variants", "Single and multi-page", "Must not be password-protected; cannot be encrypted"),
    ("JPEG", "All quality levels", "Single page", "Handles poor quality, handwritten, smartphone photos"),
    ("PNG",  "Non-animated only",  "Single page", "Animated PNGs rejected"),
    ("WebP", "Standard",           "Single page", "—"),
    ("TIFF", "All variants",       "Single or multi-page", "—"),
    ("HEIC", "Apple HEIF variant", "Single page", "iPhone default format"),
    ("HEIF", "Standard HEIF",      "Single page", "—"),
]

CUSTOM_MODEL_STEPS = [
    (1, "Go to app.mindee.com/create-model",
     "Click 'Create Custom Model'",
     "Entry point to the model builder"),
    (2, "Choose creation method",
     "Option A: Browse Catalog → select nearest template (e.g. Invoice) → auto-generates schema\nOption B: Upload sample doc → AI Agent analyses it → proposes data schema\nOption C: Type document description → AI Agent generates schema from text",
     "No ML expertise needed; AI scaffolds initial schema"),
    (3, "Review and edit Data Schema",
     "Add / remove / rename fields in the Data Schema tab.\nChoose field type (Text, Number, Date, Boolean, Classification, Nested Object, Object Detection, Barcode).\nAdd description and guidelines per field.",
     "25 field maximum recommended; global guidelines available"),
    (4, "Enable optional features",
     "Toggle: Confidence Score + Accuracy Boost | Polygons | Raw Text OCR | Continuous Learning (RAG)",
     "Per-model defaults; can also be overridden per API call"),
    (5, "Use Live Test",
     "Upload a real document in the browser. See extracted fields, confidence levels, polygon overlays, and raw JSON response.",
     "Validate schema before writing any integration code"),
    (6, "Get Model ID + API Key",
     "Copy model_id from model settings.\nCreate / copy API key from Settings → API Keys (format: md_xxxx).",
     "Model ID needed for SDK/API calls; API key authenticates all requests"),
    (7, "Integrate",
     "SDK: client.enqueue_and_parse(model_id, file)\nDirect API: POST /v2/inferences/enqueue with {model_id, file}\nNo-code: Mindee V2 module in Make / n8n / Zapier",
     "Asynchronous processing; SDK handles polling automatically"),
    (8, "Improve over time",
     "Review Low/Medium confidence extractions in platform → provide corrections → RAG feedback loop improves model durably.",
     "Continuous learning; no re-training from scratch needed"),
]

# ═══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def build_overview(wb):
    ws = wb.create_sheet("Platform Overview")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["navy"]
    r = title_block(ws, "Mindee Platform — Complete Capabilities Overview",
        "Researched from docs.mindee.com and app.mindee.com | Account: drop2ramu@gmail.com | API Key: aifactory",
        C["navy"])
    cols = ["Category", "Feature / Module", "What It Does", "Key Benefit", "Availability"]
    hdr_row(ws, r, cols, [C["navy"]]*5)
    r += 1
    cat_colors = {
        "Core": C["sky"], "Utility": C["mint"], "Optional": C["lav"],
        "Integration": C["peach"], "Platform": C["lemon"], "API": C["ice"],
    }
    for i, row in enumerate(PLATFORM_OVERVIEW):
        bg = cat_colors.get(row[0], C["lgrey"])
        data_row(ws, r, list(row), bg=bg if i%2==0 else C["white"], wrap=True, center_cols={1,5})
        ws.cell(r, 1).font = font(True, 10, C["navy"])
        ws.row_dimensions[r].height = 42
        r += 1
    set_widths(ws, [14, 24, 52, 42, 14])
    freeze(ws, "A4")

def build_models(wb):
    ws = wb.create_sheet("All Models")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["blue"]
    r = title_block(ws, "Mindee — All Models: Fields & Capabilities",
        "14 pre-built extraction models + 4 document utility tools (Crop/Split/Classify/OCR). Field counts from app.mindee.com.",
        C["blue"])
    cols = ["Category", "Model Name", "API Endpoint Slug", "Document Types", "Field Count", "All Extractable Fields"]
    hdr_row(ws, r, cols, [C["blue"]]*6)
    r += 1
    cat_bg = {
        "Finance":           C["mint"],
        "Identity":          C["sky"],
        "Healthcare":        C["lav"],
        "HR":                C["lemon"],
        "Logistics":         C["peach"],
        "Labels":            C["ice"],
        "Utility — Crop":    C["pink"],
        "Utility — Split":   C["pink"],
        "Utility — OCR":     C["pink"],
        "Utility — Classify":C["pink"],
    }
    for i, row in enumerate(MODELS_ALL):
        bg = cat_bg.get(row[0], C["lgrey"])
        data_row(ws, r, list(row), bg=bg if i%2==0 else C["white"], wrap=True, center_cols={1,3,5})
        ws.cell(r, 2).font = font(True, 10, C["blue"])
        ws.row_dimensions[r].height = 60
        r += 1
    set_widths(ws, [20, 22, 28, 28, 11, 68])
    freeze(ws, "A4")

def build_extraction_detail(wb):
    ws = wb.create_sheet("Extraction Deep Dive")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["teal"]
    r = title_block(ws, "Extraction Models — Deep Dive by Model",
        "Detailed field-by-field breakdown for 6 core pre-built extraction models. Field counts verified from app.mindee.com.",
        C["teal"])

    models_detail = [
        ("Invoice  (26 fields — Supplier Name + 25 more)", C["mint"], [
            ("supplier_name",          "Text",    "Supplier company name"),
            ("invoice_number",         "Text",    "Invoice identifier"),
            ("date",                   "Date",    "Invoice issue date (YYYY-MM-DD)"),
            ("due_date",               "Date",    "Payment due date"),
            ("payment_date",           "Date",    "Date payment was made"),
            ("document_type",          "Class",   "invoice | payslip | quote | purchase_order | statement | receipt | credit_note | other_financial"),
            ("total_amount",           "Number",  "Final total including all taxes"),
            ("total_net",              "Number",  "Total before taxes"),
            ("total_tax",              "Number",  "Total tax amount"),
            ("taxes[]",                "Array",   "rate (Number) + base (Number) + amount (Number)"),
            ("line_items[]",           "Array",   "description + quantity + unit_price + total_price + product_code + tax_amount + tax_rate + unit_measure"),
            ("supplier_address",       "Object",  "address, street_number, street_name, po_box, address_complement, city, postal_code, state, country"),
            ("customer_address",       "Object",  "Same subfields as supplier_address"),
            ("billing_address",        "Object",  "Same subfields as supplier_address"),
            ("shipping_address",       "Object",  "Same subfields as supplier_address"),
            ("customer_name",          "Text",    "Customer company/person name"),
            ("customer_id",            "Text",    "Customer identifier"),
            ("supplier_phone_number",  "Text",    "Supplier phone"),
            ("supplier_email",         "Text",    "Supplier email address"),
            ("supplier_website",       "Text",    "Supplier website URL"),
            ("po_number",              "Text",    "Purchase order number"),
            ("reference_numbers[]",    "Array",   "Additional reference strings"),
            ("supplier_payment_details[]", "Array","IBAN + SWIFT + account_number + routing_number"),
            ("customer_company_registration[]", "Array", "number + type (VAT/SIRET/SIREN/TIN/RFC/…)"),
            ("supplier_company_registration[]","Array", "number + type (VAT/SIRET/SIREN/TIN/RFC/…)"),
            ("locale",                 "Object",  "language (ISO 639-1) + country (ISO 3166-1) + currency (ISO 4217)"),
        ]),
        ("Receipt  (17 fields — Supplier Name + 16 more)", C["sky"], [
            ("supplier_name",                  "Text",   "Store/restaurant name"),
            ("supplier_address",               "Text",   "Supplier address string"),
            ("supplier_phone_number",          "Text",   "Supplier phone"),
            ("supplier_company_registration",  "Object", "number + type"),
            ("receipt_number",                 "Text",   "Receipt identifier"),
            ("date",                           "Date",   "Transaction date"),
            ("time",                           "Text",   "Transaction time"),
            ("document_type",                  "Class",  "expense_receipt | credit_card_receipt"),
            ("total_amount",                   "Number", "Final total including taxes/discounts"),
            ("total_net",                      "Number", "Pre-tax total"),
            ("total_tax",                      "Number", "Total tax amount"),
            ("tips_gratuity",                  "Number", "Tip amount"),
            ("taxes[]",                        "Array",  "rate + base + amount"),
            ("line_items[]",                   "Array",  "description + quantity + unit_price + total_price"),
            ("purchase_category",              "Class",  "food | gasoline | parking | toll | accommodation | transport | telecom | software | shopping | energy | miscellaneous"),
            ("purchase_subcategory",           "Class",  "restaurant | delivery | train | public | taxi | car_rental | plane | micromobility | office_supplies | electronics | cultural | groceries | other"),
            ("locale",                         "Object", "language + country + currency (ISO standards)"),
        ]),
        ("Passport  (12 fields — Given Names + 11 more)", C["lav"], [
            ("given_names",           "Text",  "Holder's first names"),
            ("surnames",              "Text",  "Holder's last names"),
            ("date_of_birth",         "Date",  "YYYY-MM-DD"),
            ("place_of_birth",        "Text",  "Birth location"),
            ("passport_number",       "Text",  "Document identifier"),
            ("issuing_country",       "Text",  "Country that issued the passport"),
            ("nationality",           "Text",  "Holder's citizenship"),
            ("date_of_issue",         "Date",  "YYYY-MM-DD"),
            ("date_of_expiry",        "Date",  "YYYY-MM-DD"),
            ("sex",                   "Class", "Male | Female | Other"),
            ("mrz_line_1",            "Text",  "Machine Readable Zone line 1"),
            ("mrz_line_2",            "Text",  "Machine Readable Zone line 2"),
            ("— India extras —",      "—",     "Additional fields extracted from Indian passports only"),
            ("legal_guardian",        "Text",  "India only"),
            ("spouse_name",           "Text",  "India only"),
            ("mother_name",           "Text",  "India only"),
            ("prior_passport_number", "Text",  "India only — previous passport number"),
            ("file_number",           "Text",  "India only — government file number"),
            ("address",               "Text",  "India only — 3-line address"),
        ]),
        ("US Healthcare Card  (14 fields)", C["peach"], [
            ("company_name",     "Text",    "Health plan provider company"),
            ("plan_name",        "Text",    "Healthcare plan name"),
            ("member_name",      "Text",    "Covered individual"),
            ("member_id",        "Text",    "Unique member identifier"),
            ("issuer_80840",     "Text",    "Issuing organisation"),
            ("group_number",     "Text",    "Group plan identifier"),
            ("payer_id",         "Text",    "Unique payer system ID"),
            ("dependents[]",     "Array",   "List of covered dependents"),
            ("rx_bin",           "Text",    "Pharmacy BIN for prescription coverage"),
            ("rx_id",            "Text",    "Prescription ID"),
            ("rx_grp",           "Text",    "Prescription group number"),
            ("rx_pcn",           "Text",    "Prescription PCN number"),
            ("copayments[]",     "Array",   "service_name (primary_care|ER|urgent_care|specialist|office_visit|prescription) + service_fees (Number)"),
            ("enrollment_date",  "Date",    "Plan enrollment date"),
        ]),
        ("Resume / CV  (13 fields — Name + Address + 11 more)", C["lemon"], [
            ("name",                    "Text",   "Full candidate name"),
            ("address",                 "Text",   "Current address"),
            ("phone_number",            "Text",   "Contact phone"),
            ("email",                   "Text",   "Contact email"),
            ("linkedin_profile",        "Text",   "LinkedIn URL"),
            ("education[]",             "Array",  "school_name + degree + dates_attended + gpa + relevant_coursework"),
            ("professional_history[]",  "Array",  "company_name + job_title + dates_employed + responsibilities"),
            ("skills[]",                "Array",  "Skill strings"),
            ("languages[]",             "Array",  "Language name + proficiency level"),
            ("projects[]",              "Array",  "name + description + technologies"),
            ("awards_certifications[]", "Array",  "name + date_received"),
            ("summary",                 "Text",   "Objective/summary statement"),
            ("candidate_photo",         "ObjDet", "Photo location polygon"),
        ]),
        ("Driver License  (14 fields — First Name + 13 more)", C["ice"], [
            ("first_name",        "Text",  "First name"),
            ("last_name",         "Text",  "Last name"),
            ("date_of_birth",     "Date",  "YYYY-MM-DD"),
            ("sex",               "Class", "M | F | Other"),
            ("document_id",       "Text",  "Licence number"),
            ("issued_date",       "Date",  "Issue date"),
            ("expiry_date",       "Date",  "Expiry date"),
            ("country_code",      "Text",  "ISO country code"),
            ("issuing_authority", "Text",  "Issuing authority"),
            ("mrz",               "Text",  "Machine-readable zone number"),
            ("category",          "Text",  "EU driving categories (A, B, C, D…)"),
            ("street",            "Text",  "Street address"),
            ("city",              "Text",  "City"),
            ("state",             "Text",  "State/region"),
        ]),
    ]

    for model_name, model_color, fields in models_detail:
        r = section_hdr(ws, r, f"  {model_name} Model Fields", model_color)
        hdr_row(ws, r, ["Field Name", "Type", "Description / Values"], [model_color]*3)
        r += 1
        for j, (fname, ftype, fdesc) in enumerate(fields):
            bg = C["lgrey"] if j%2==0 else C["white"]
            data_row(ws, r, [fname, ftype, fdesc], bg=bg, wrap=True, center_cols={2})
            ws.cell(r, 1).font = font(True, 9, C["teal"])
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1

    set_widths(ws, [32, 12, 70])
    freeze(ws, "A4")

def build_field_types(wb):
    ws = wb.create_sheet("Field Types & Schema")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["purple"]
    r = title_block(ws, "Data Schema — Field Types & Configuration Rules",
        "8 field types available when building custom models. Applies to all Extraction, Split, Crop, and Classification models.",
        C["purple"])

    cols = ["Field Type", "Description", "Example Fields", "Output Format", "Array Support", "Notes"]
    hdr_row(ws, r, cols, [C["purple"]]*6)
    r += 1
    type_colors = {
        "Text": C["sky"], "Number": C["mint"], "Date": C["lav"],
        "Boolean": C["lemon"], "Classification": C["peach"],
        "Nested Object": C["ice"], "Object Detection": C["pink"], "Barcode": C["sand"],
    }
    for i, row in enumerate(FIELD_TYPES):
        bg = type_colors.get(row[0], C["lgrey"])
        data_row(ws, r, list(row), bg=bg, wrap=True, center_cols={4})
        ws.cell(r, 1).font = font(True, 10, C["purple"])
        ws.row_dimensions[r].height = 42
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Schema Configuration Rules", C["purple"])
    hdr_row(ws, r, ["Rule", "Detail"], [C["purple"]]*2)
    r += 1
    for i, (rule, detail) in enumerate(FIELD_RULES):
        data_row(ws, r, [rule, detail], bg=C["lgrey"] if i%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["purple"])
        ws.row_dimensions[r].height = 28
        r += 1

    set_widths(ws, [20, 34, 30, 30, 14, 34])
    freeze(ws, "A4")

def build_utility_models(wb):
    ws = wb.create_sheet("Utility Models")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["green"]
    r = title_block(ws, "Utility Models — Crop / Split / Classification / OCR",
        "Four beta utility models for document preprocessing and routing. All are custom — no pre-built templates.",
        C["green"])

    sections = [
        ("Auto-Crop Model", C["teal"], [
            ("Purpose",           "Detects and localises individual document regions on a single scanned page with multiple documents."),
            ("Output per region", "page_index (0-based) + bounding_box (normalised polygon, clockwise from top-left) + classification (user-defined class)"),
            ("Class names",       "Returned exactly as defined on platform, including spaces and capitalisation. Changes take effect immediately."),
            ("Model creation",    "Custom only — each model gets its own unique model ID. Define classes for each document type you expect."),
            ("Use cases",         "Batch scans with multiple IDs on one page; mixed receipts/invoices in one capture; multi-card scans."),
            ("Chaining",          "Output polygons can feed directly into an Extraction model via the 'extraction-model-chaining' feature."),
            ("File limits",       "Same as standard API: 100 MB, 200 pages, PDF/JPEG/PNG/WebP/TIFF/HEIC/HEIF."),
            ("Coordinate format", "Normalised floats [0.0–1.0]. (0,0)=top-left, (1,1)=bottom-right. Clockwise point order. Index 0 = top-left corner."),
        ]),
        ("Auto-Split Model", C["orange"], [
            ("Purpose",           "Breaks a multi-page PDF into separate logical documents by detecting where one document ends and the next begins."),
            ("Output per segment","page_start (0-based) + page_end (0-based) + category (user-defined class)"),
            ("Class names",       "Returned exactly as defined. Optional 'OTHER' class captures unmatched document types."),
            ("Use cases",         "Multiple invoices in one PDF; ID card front+back as separate segments; mixed invoice/receipt/statement bundles; regional language variants."),
            ("Model creation",    "Custom only — define your document classes (e.g. INVOICE, RECEIPT, DRIVER_LICENSE, OTHER)."),
            ("Chaining",          "Split ranges can chain into Extraction models to automatically extract fields from each detected segment."),
            ("No pre-built templates", "Users define all classes from scratch. AI Agent assists with schema if doc samples provided."),
            ("Class changes",     "Renaming a class in the platform immediately changes what the API returns — no model redeployment needed."),
        ]),
        ("Classification Model", C["indigo"], [
            ("Purpose",           "Assigns a single class to an entire document (single or multi-page) — examines all pages before deciding."),
            ("Output",            "predicted_class (one of your user-defined classes) + confidence score"),
            ("Class names",       "Returned exactly as defined on platform, including spaces and capitalisation."),
            ("Use cases",         "Routing documents to correct department; language/region identification; separating financial vs identity documents."),
            ("Optional OTHER",    "Add an 'OTHER' class to catch document types outside your defined set without errors."),
            ("Model creation",    "Custom only — define classes (e.g. INVOICES, IDENTITY DOCUMENTS, CONTRACTS)."),
            ("Multi-page",        "All pages analysed together before classification — handles documents where doc type only clear from later pages."),
            ("Chaining",          "Classification output can route to the appropriate Extraction model automatically."),
        ]),
        ("Raw Text OCR Model", C["brown"], [
            ("Purpose",           "Extracts all text from a document at full-page and word level with normalised coordinate positions."),
            ("Output per page",   "full_text (entire page as one string) + words[] (array of {text, polygon, page_index})"),
            ("Polygon format",    "Normalised [0.0–1.0] coordinates per word, clockwise from top-left."),
            ("Language support",  "Almost all modern languages. Excludes: ancient scripts (cuneiform, hieroglyphics), Blackfoot, Cherokee, Inuktitut."),
            ("Use cases",         "RAG pipeline text ingestion; search indexing; NLP preprocessing; document archiving."),
            ("vs Extraction",     "OCR returns raw text only — no field parsing. Use alongside Extraction models when you need both."),
            ("Model creation",    "Custom — define model once, apply to any document type."),
            ("Optional features", "Polygons, confidence scores, and continuous learning all apply to OCR models too."),
        ]),
    ]

    for section_name, sec_color, rows in sections:
        r = section_hdr(ws, r, f"  {section_name}", sec_color)
        hdr_row(ws, r, ["Aspect", "Detail"], [sec_color]*2)
        r += 1
        for j, (aspect, detail) in enumerate(rows):
            data_row(ws, r, [aspect, detail], bg=C["lgrey"] if j%2==0 else C["white"], wrap=True)
            ws.cell(r, 1).font = font(True, 10, sec_color)
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    set_widths(ws, [26, 88])
    freeze(ws, "A4")

def build_optional_features(wb):
    ws = wb.create_sheet("Optional Features")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["orange"]
    r = title_block(ws, "Optional Features — Confidence, Polygons, Raw Text, RAG",
        "Four opt-in features available per model (platform defaults) or overridden per API call.",
        C["orange"])

    r = section_hdr(ws, r, "  Confidence Score + Accuracy Boost", C["orange"])
    conf_intro = [
        ("How it works",    "Ensemble of multiple independently-trained models. Agreement between models drives confidence level. Dedicated arbitration model checks structural + semantic coherence."),
        ("Score levels",    "Certain (Blue) | High (Green) | Medium (Orange) | Low (Red)"),
        ("Automation use",  "Certain + High → safe for full automation. Medium → optional review. Low → manual review queue + RAG feedback signal."),
        ("Accuracy Boost",  "Combines ensemble predictions via consensus algorithms, selecting the most reliable result across all models."),
        ("RAG integration", "Low-confidence extractions automatically become RAG feedback signals, improving the model durably over time."),
        ("Coverage",        "Works with all field types including nested objects, arrays (line items, taxes), and object detection."),
        ("Latency",         "Minimal overhead — models run in parallel."),
        ("Activation",      "Toggle in model settings (platform default) or pass feature flag per API call for granular control."),
    ]
    hdr_row(ws, r, ["Aspect", "Detail"], [C["orange"]]*2)
    r += 1
    for j, (k, v) in enumerate(conf_intro):
        data_row(ws, r, [k, v], bg=C["peach"] if j%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["orange"])
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Confidence Level Reference", C["orange"])
    hdr_row(ws, r, ["Level", "Colour", "Meaning", "Recommended Action", "Technical Detail"], [C["orange"]]*5)
    r += 1
    level_fills = {"Certain": "1565C0", "High": "2E7D32", "Medium": "E65100", "Low": "B71C1C"}
    for j, row in enumerate(CONFIDENCE_LEVELS):
        data_row(ws, r, list(row), bg=C["lgrey"] if j%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(True, 10, "FFFFFF")
        ws.cell(r, 1).fill = fill(level_fills.get(row[0], C["navy"]))
        ws.row_dimensions[r].height = 38
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Polygons / Bounding Boxes", C["teal"])
    poly_rows = [
        ("Coordinate system", "Normalised floats 0.0–1.0. Top-left = (0,0). Bottom-right = (1,1)."),
        ("Format",            "Array of points forming a closed polygon; clockwise order; index 0 = top-left corner."),
        ("Response path",     "response.inference.result.fields.<name>.locations[i].polygon"),
        ("Page index",        "response.inference.result.fields.<name>.locations[i].page (0-based)"),
        ("Centroid helper",   ".polygon.centroid — geometric centre of the polygon, useful for UI overlay positioning."),
        ("Multiple locations","A field may have multiple location objects if it appears on multiple pages."),
        ("Use cases",         "Visual overlay highlighting extracted fields; debugging extraction quality; UX field correction UI."),
        ("Activation",        "Toggle in model settings or per API call. Returns null when inactive."),
    ]
    hdr_row(ws, r, ["Aspect", "Detail"], [C["teal"]]*2)
    r += 1
    for j, (k, v) in enumerate(poly_rows):
        data_row(ws, r, [k, v], bg=C["ice"] if j%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["teal"])
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Raw Text Full OCR  +  Continuous Learning (RAG)", C["green"])
    rag_rows = [
        ("Raw Text — What",       "Adds complete page text to every API response alongside structured fields."),
        ("Raw Text — Path",       "response.inference.result.options.raw_text (list of strings, one per page)"),
        ("Raw Text — Words",      "response.inference.result.options.words — array of {text, polygon, page_index}"),
        ("Raw Text — Benefit",    "No second OCR API call needed; feed downstream NLP/RAG in one request."),
        ("RAG — What",            "Corrections to Low/Medium confidence extractions fed back into the model via Retrieval-Augmented Generation."),
        ("RAG — How",             "Platform UI: review problematic extractions → provide correct values → model improves durably without full retraining."),
        ("RAG — Benefit",         "Accuracy compounds over time; domain-specific documents improve fastest (medical, legal, logistics)."),
        ("RAG — Activation",      "Enable in model settings. Low-confidence fields automatically become candidate feedback items in the review queue."),
    ]
    hdr_row(ws, r, ["Aspect", "Detail"], [C["green"]]*2)
    r += 1
    for j, (k, v) in enumerate(rag_rows):
        data_row(ws, r, [k, v], bg=C["mint"] if j%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["green"])
        ws.row_dimensions[r].height = 26
        r += 1

    set_widths(ws, [24, 90])
    freeze(ws, "A4")

def build_api_integration(wb):
    ws = wb.create_sheet("API & Integration")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["indigo"]
    r = title_block(ws, "API Reference — Endpoints, Response Format, Limits",
        "Async RESTful API. All file processing routes are asynchronous. SDKs abstract polling automatically.",
        C["indigo"])

    r = section_hdr(ws, r, "  API Endpoints", C["indigo"])
    hdr_row(ws, r, ["Method", "Path", "Purpose", "Auth Header", "Request Body / Params", "Response"],
            [C["indigo"]]*6)
    r += 1
    for j, row in enumerate(API_ENDPOINTS):
        data_row(ws, r, list(row), bg=C["lgrey"] if j%2==0 else C["white"], wrap=True, center_cols={1})
        ws.cell(r, 1).font = font(True, 10, "FFFFFF")
        ws.cell(r, 1).fill = fill(C["green"] if row[0]=="GET" else C["orange"])
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1
    r = section_hdr(ws, r, "  API Response JSON Structure", C["indigo"])
    hdr_row(ws, r, ["JSON Path", "Type", "Description"], [C["indigo"]]*3)
    r += 1
    for j, row in enumerate(RESPONSE_STRUCTURE):
        data_row(ws, r, list(row), bg=C["lav"] if j%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(False, 9, C["indigo"], name="Courier New")
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Rate Limits & File Constraints", C["red"])
    hdr_row(ws, r, ["Limit Type", "Value", "HTTP Error Code", "Applies To"], [C["red"]]*4)
    r += 1
    for j, row in enumerate(RATE_LIMITS):
        data_row(ws, r, list(row), bg=C["pink"] if j%2==0 else C["white"], wrap=True)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Supported File Formats", C["brown"])
    hdr_row(ws, r, ["Format", "Variants Supported", "Page Support", "Notes"], [C["brown"]]*4)
    r += 1
    for j, row in enumerate(FILE_FORMATS):
        data_row(ws, r, list(row), bg=C["sand"] if j%2==0 else C["white"], wrap=True)
        ws.row_dimensions[r].height = 22
        r += 1

    set_widths(ws, [14, 34, 28, 28, 36, 34])
    freeze(ws, "A4")

def build_sdks(wb):
    ws = wb.create_sheet("SDKs & No-Code")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["cyan"]
    r = title_block(ws, "SDKs & No-Code Integrations — 6 Languages + 3 Platforms",
        "Official SDKs handle file upload, async polling, response parsing, and error management automatically.",
        C["cyan"])

    r = section_hdr(ws, r, "  Official SDK Reference", C["cyan"])
    hdr_row(ws, r, ["Language", "Install Command", "Key Code Pattern", "Notes"], [C["cyan"]]*4)
    r += 1
    sdk_colors = [C["sky"], C["mint"], C["lav"], C["peach"], C["lemon"], C["ice"], C["lgrey"]]
    for j, row in enumerate(SDK_DETAILS):
        data_row(ws, r, list(row), bg=sdk_colors[j % len(sdk_colors)], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["cyan"])
        ws.cell(r, 3).font = font(False, 9, C["black"], name="Courier New")
        ws.row_dimensions[r].height = 55
        r += 1

    r += 1
    r = section_hdr(ws, r, "  No-Code Platform Integrations", C["green"])
    hdr_row(ws, r, ["Platform", "Setup Steps", "Trigger Sources", "Extract Action", "Output Destinations", "Notes"],
            [C["green"]]*6)
    r += 1
    nc_colors = [C["mint"], C["lemon"], C["peach"]]
    for j, row in enumerate(NOCODE_INTEGRATIONS):
        data_row(ws, r, list(row), bg=nc_colors[j], wrap=True)
        ws.cell(r, 1).font = font(True, 11, C["green"])
        ws.row_dimensions[r].height = 65
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Webhook Technical Reference", C["orange"])
    hdr_row(ws, r, ["Aspect", "Detail"], [C["orange"]]*2)
    r += 1
    for j, (aspect, detail) in enumerate(WEBHOOK_DETAILS):
        data_row(ws, r, [aspect, detail], bg=C["peach"] if j%2==0 else C["white"], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["orange"])
        ws.row_dimensions[r].height = 26
        r += 1

    set_widths(ws, [18, 36, 28, 28, 28, 32])
    freeze(ws, "A4")

def build_custom_model(wb):
    ws = wb.create_sheet("Custom Model Guide")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["brown"]
    r = title_block(ws, "Custom Model Builder — Step-by-Step Guide",
        "Build a bespoke extraction model at app.mindee.com with your API key (aifactory). No ML expertise needed.",
        C["brown"])

    r = section_hdr(ws, r, "  8-Step Process to Create and Deploy a Custom Model", C["brown"])
    hdr_row(ws, r, ["Step", "Where", "Action", "Technical Detail"], [C["brown"]]*4)
    r += 1
    step_colors = [C["sky"], C["mint"], C["lav"], C["peach"], C["lemon"], C["ice"], C["pink"], C["sand"]]
    for j, (step, where, action, detail) in enumerate(CUSTOM_MODEL_STEPS):
        data_row(ws, r, [step, where, action, detail], bg=step_colors[j], wrap=True, center_cols={1})
        ws.cell(r, 1).font = font(True, 14, C["brown"])
        ws.cell(r, 1).alignment = align("center", "center")
        ws.row_dimensions[r].height = 52
        r += 1

    r += 1
    r = section_hdr(ws, r, "  Catalog vs Custom — Decision Guide", C["teal"])
    hdr_row(ws, r, ["Approach", "When to Use", "Starting Point", "Time to First Result", "Customisation"],
            [C["teal"]]*5)
    r += 1
    approaches = [
        ("Catalog Template",
         "Your document matches one of the 17 pre-built types (invoice, receipt, passport, etc.)",
         "Browse Catalog → select template → model ready instantly",
         "< 1 minute",
         "Add/remove/rename fields in Data Schema tab"),
        ("AI From Sample",
         "You have a sample document but no matching catalog template",
         "Upload sample → AI Agent proposes schema → review and adjust",
         "2–5 minutes",
         "Full schema control; AI suggestions as starting point"),
        ("AI From Description",
         "You can describe the document type but don't have a sample",
         "Type document description → AI Agent generates schema",
         "1–3 minutes",
         "Full schema control; test with Live Test before integrating"),
        ("From Scratch",
         "You need precise control over every field from the start",
         "Click Create Custom Model → manually add each field",
         "10–30 minutes depending on complexity",
         "Full schema control from first field"),
    ]
    approach_colors = [C["mint"], C["sky"], C["lav"], C["peach"]]
    for j, row in enumerate(approaches):
        data_row(ws, r, list(row), bg=approach_colors[j], wrap=True)
        ws.cell(r, 1).font = font(True, 10, C["teal"])
        ws.row_dimensions[r].height = 46
        r += 1

    set_widths(ws, [6, 28, 28, 28, 24])
    freeze(ws, "A4")

def build_quick_ref(wb):
    ws = wb.create_sheet("Quick Reference")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["dgrey"]
    r = title_block(ws, "Quick Reference — Your Account Details + Key API Patterns",
        "Account: drop2ramu@gmail.com | API Key Name: aifactory | Key: md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk",
        C["dgrey"])

    account = [
        ("Account Email",  "drop2ramu@gmail.com"),
        ("API Key Name",   "aifactory"),
        ("API Key Value",  "md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk"),
        ("API Version",    "V2  (app.mindee.com — NOT legacy platform.mindee.com)"),
        ("Base URL",       "https://api.mindee.net/v2/"),
        ("Docs URL",       "https://docs.mindee.com/"),
        ("App URL",        "https://app.mindee.com/"),
        ("Create Model",   "https://app.mindee.com/create-model"),
    ]
    r = section_hdr(ws, r, "  Account & Connection Details", C["navy"])
    for j, (k, v) in enumerate(account):
        data_row(ws, r, [k, v], bg=C["sky"] if j%2==0 else C["white"], wrap=False)
        ws.cell(r, 1).font = font(True, 10, C["navy"])
        ws.cell(r, 2).font = font(False, 10, name="Courier New")
        ws.row_dimensions[r].height = 18
        r += 1

    snippets = [
        ("Python SDK — Invoice", C["mint"],
         'pip install mindee\n\nfrom mindee import Client, product\n\nclient = Client(api_key="md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk")\n\n# Submit and wait for result (SDK polls automatically)\nresult = client.enqueue_and_parse(product.InvoiceV4, "./invoice.pdf")\n\n# Access fields\nprint(result.document.inference.result.fields["supplier_name"].value)\nprint(result.document.inference.result.fields["total_amount"].value)\nprint(result.document.inference.result.fields["date"].value)\n\n# Line items (array field)\nfor item in result.document.inference.result.fields["line_items"].items:\n    print(item.fields["description"].value, item.fields["total_price"].value)'),
        ("Python SDK — Custom Model", C["sky"],
         'from mindee import Client\nfrom mindee.product import CustomV1  # use CustomV1 for any custom model\n\nclient = Client(api_key="md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk")\n\n# Replace YOUR_MODEL_ID with the model ID from app.mindee.com\nresult = client.enqueue_and_parse(\n    CustomV1,\n    "./document.pdf",\n    endpoint_name="YOUR_MODEL_ID",\n)\n\nfor name, field in result.document.inference.result.fields.items():\n    print(f"{name}: {field.value}  (confidence: {field.confidence})")'),
        ("Direct REST API — Enqueue + Poll", C["lav"],
         '# Step 1 — Submit file\ncurl -X POST "https://api.mindee.net/v2/inferences/enqueue" \\\n  -H "Authorization: Token md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk" \\\n  -F "file=@invoice.pdf" \\\n  -F "model_id=mindee/invoices/v4"\n# Returns: {"job_id": "xxx", "status": "enqueued"}\n\n# Step 2 — Wait 3 seconds, then poll\nsleep 3\ncurl "https://api.mindee.net/v2/jobs/JOB_ID_HERE" \\\n  -H "Authorization: Token md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk"\n# HTTP 200 = still processing | HTTP 302 = done (use result_url)\n\n# Step 3 — Get result\ncurl "https://api.mindee.net/v2/inferences/INFERENCE_ID_HERE" \\\n  -H "Authorization: Token md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk"'),
        ("Make.com Setup", C["peach"],
         'Steps:\n1. In Make.com, search "Mindee V2 verified" (NOT community apps)\n2. Select Action: "Extract Document Data"\n3. Create Connection:\n   - Name:    MindeeV2-aifactory\n   - API Key: md_9bHNrv8-gFT7x0TjQaZcbh6ZrjgYICzkP4XDVzpJdrk\n4. Click "Search Model" → type model name (e.g. "Invoice")\n5. Link a file input module (Gmail, Drive, Dropbox, etc.)\n   IMPORTANT: Do NOT fill the "File" field manually — it auto-connects\n6. Map output fields to downstream modules (Sheets, CRM, Slack, etc.)'),
    ]

    r += 1
    for title, color, code in snippets:
        r = section_hdr(ws, r, f"  {title}", color, span=2)
        cell = ws.cell(r, 1, code)
        cell.font      = Font(name="Courier New", size=9, color=C["black"])
        cell.fill      = fill("F8F8F8")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border    = border()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lines = code.count("\n") + 1
        ws.row_dimensions[r].height = max(14 * lines, 60)
        r += 2

    set_widths(ws, [55, 55])
    freeze(ws, "A4")

# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    out_dir  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Mindee_Capabilities.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building Mindee_Capabilities.xlsx ...")
    build_overview(wb);         print("  [+] Platform Overview")
    build_models(wb);            print("  [+] All Models")
    build_extraction_detail(wb); print("  [+] Extraction Deep Dive")
    build_field_types(wb);       print("  [+] Field Types & Schema")
    build_utility_models(wb);    print("  [+] Utility Models (Crop/Split/Classify/OCR)")
    build_optional_features(wb); print("  [+] Optional Features")
    build_api_integration(wb);   print("  [+] API & Integration")
    build_sdks(wb);              print("  [+] SDKs & No-Code")
    build_custom_model(wb);      print("  [+] Custom Model Guide")
    build_quick_ref(wb);         print("  [+] Quick Reference")

    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"Sheets ({len(wb.sheetnames)}):")
    for s in wb.sheetnames:
        print(f"  - {s}")

if __name__ == "__main__":
    main()
